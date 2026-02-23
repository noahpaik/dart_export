"""Excel output module."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWriterError(RuntimeError):
    """Raised when excel write operation fails."""


@dataclass
class WriteSummary:
    sheet_name: str
    written_cells: int
    matched_accounts: int
    unmatched_accounts: list[tuple[str, str]]


class ExcelWriter:
    """
    Excel template writer for Track A/Track B outputs.

    Default template conventions:
      - header row for years: row 3
      - account id column: A (1)
      - account name column: B (2)
    """

    STATEMENT_ALIASES: dict[str, str] = {
        "dart_ShortTermDepositsNotClassifiedAsCashEquivalents": "ifrs-full_CurrentFinancialAssets",
        "dart_ShortTermBorrowings": "ifrs-full_ShorttermBorrowings",
        "dart_BondsIssued": "ifrs-full_BondsIssued",
        "dart_TotalSellingGeneralAdministrativeExpenses": "dart_TotalSGA",
        "ifrs-full_TradeReceivables": "ifrs-full_TradeAndOtherCurrentReceivables",
        "ifrs-full_TradePayables": "ifrs-full_TradeAndOtherCurrentPayables",
    }

    def __init__(
        self,
        template_path: str,
        year_header_row: int = 3,
        account_id_col: int = 1,
        account_name_col: int = 2,
        create_if_missing: bool = False,
    ):
        self.template_path = Path(template_path)
        self.year_header_row = year_header_row
        self.account_id_col = account_id_col
        self.account_name_col = account_name_col
        self.write_logs: list[WriteSummary] = []

        if self.template_path.exists():
            self.wb = load_workbook(self.template_path)
        else:
            if not create_if_missing:
                raise ExcelWriterError(f"엑셀 템플릿 파일이 없습니다: {self.template_path}")
            self.wb = self._create_default_workbook()

    def _create_default_workbook(self) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "BS"
        for sheet_name in ("IS", "CF", "판관비", "부문별매출"):
            wb.create_sheet(title=sheet_name)
        return wb

    def write_balance_sheet(self, df: pd.DataFrame, sheet_name: str = "BS") -> WriteSummary:
        return self._write_statement(df, sheet_name)

    def write_income_statement(self, df: pd.DataFrame, sheet_name: str = "IS") -> WriteSummary:
        return self._write_statement(df, sheet_name)

    def write_cash_flow(self, df: pd.DataFrame, sheet_name: str = "CF") -> WriteSummary:
        return self._write_statement(df, sheet_name)

    def _write_statement(self, df: pd.DataFrame, sheet_name: str) -> WriteSummary:
        if df is None or df.empty:
            summary = WriteSummary(sheet_name=sheet_name, written_cells=0, matched_accounts=0, unmatched_accounts=[])
            self.write_logs.append(summary)
            return summary

        ws = self._get_sheet(sheet_name)
        required_columns = {"계정ID"}
        if not required_columns.issubset(set(df.columns)):
            raise ExcelWriterError(f"{sheet_name}: 필수 컬럼 누락 {required_columns} (columns={list(df.columns)})")

        year_columns = [
            col
            for col in df.columns
            if col not in {"계정ID", "계정명", "재무제표구분", "재무제표명", "계정상세"}
        ]
        if not year_columns:
            raise ExcelWriterError(f"{sheet_name}: 연도 컬럼이 없습니다.")

        col_map = self._ensure_year_headers(ws, [str(col) for col in year_columns])
        id_row_map = self._read_id_row_map(ws)
        if not id_row_map:
            id_row_map = self._bootstrap_statement_rows(ws, df)

        unmatched: list[tuple[str, str]] = []
        matched_accounts = 0
        written_cells = 0

        for _, row in df.iterrows():
            account_id = str(row.get("계정ID", "")).strip()
            account_name = str(row.get("계정명", "")).strip()
            if not account_id:
                continue

            target_row = self._resolve_target_row(account_id, id_row_map)
            if target_row is None:
                unmatched.append((account_id, account_name))
                continue

            matched_accounts += 1
            for year in year_columns:
                if year not in col_map:
                    continue
                value = self._to_number_or_none(row.get(year))
                if value is None:
                    continue
                ws.cell(row=target_row, column=col_map[year], value=value)
                written_cells += 1

        summary = WriteSummary(
            sheet_name=sheet_name,
            written_cells=written_cells,
            matched_accounts=matched_accounts,
            unmatched_accounts=unmatched,
        )
        self.write_logs.append(summary)
        return summary

    def _resolve_target_row(self, account_id: str, id_row_map: dict[str, int]) -> int | None:
        if account_id in id_row_map:
            return id_row_map[account_id]

        alias = self.STATEMENT_ALIASES.get(account_id)
        if alias and alias in id_row_map:
            return id_row_map[alias]

        for src, dst in self.STATEMENT_ALIASES.items():
            if dst == account_id and src in id_row_map:
                return id_row_map[src]
        return None

    def write_sga_detail(
        self,
        data: dict[str, dict[str, Any]] | pd.DataFrame,
        sheet_name: str = "판관비",
    ) -> WriteSummary:
        ws = self._get_sheet(sheet_name)
        normalized = self._normalize_named_year_values(data)
        years = sorted({year for yearly in normalized.values() for year in yearly.keys()})
        col_map = self._ensure_year_headers(ws, years)
        row_map = self._read_account_name_row_map(ws)
        if not row_map:
            row_map = self._bootstrap_named_rows(ws, list(normalized.keys()))

        unmatched: list[tuple[str, str]] = []
        matched_accounts = 0
        written_cells = 0

        for account_name, yearly in normalized.items():
            row_idx = row_map.get(account_name)
            if row_idx is None:
                unmatched.append((account_name, account_name))
                continue
            matched_accounts += 1
            for year, value in yearly.items():
                col_idx = col_map.get(year)
                if col_idx is None:
                    continue
                number = self._to_number_or_none(value)
                if number is None:
                    continue
                ws.cell(row=row_idx, column=col_idx, value=number)
                written_cells += 1

        summary = WriteSummary(
            sheet_name=sheet_name,
            written_cells=written_cells,
            matched_accounts=matched_accounts,
            unmatched_accounts=unmatched,
        )
        self.write_logs.append(summary)
        return summary

    def write_segment_revenue(
        self,
        data: dict[str, dict[str, Any]] | pd.DataFrame,
        sheet_name: str = "부문별매출",
        start_row: int = 5,
    ) -> WriteSummary:
        ws = self._get_sheet(sheet_name)
        normalized = self._normalize_named_year_values(data)
        years = sorted({year for yearly in normalized.values() for year in yearly.keys()})
        col_map = self._ensure_year_headers(ws, years)

        written_cells = 0
        matched_accounts = 0
        for idx, (segment_name, yearly) in enumerate(normalized.items()):
            row_idx = start_row + idx
            ws.cell(row=row_idx, column=self.account_name_col, value=segment_name)
            matched_accounts += 1
            for year, value in yearly.items():
                col_idx = col_map.get(year)
                if col_idx is None:
                    continue
                number = self._to_number_or_none(value)
                if number is None:
                    continue
                ws.cell(row=row_idx, column=col_idx, value=number)
                written_cells += 1

        summary = WriteSummary(
            sheet_name=sheet_name,
            written_cells=written_cells,
            matched_accounts=matched_accounts,
            unmatched_accounts=[],
        )
        self.write_logs.append(summary)
        return summary

    def _normalize_named_year_values(
        self,
        data: dict[str, dict[str, Any]] | pd.DataFrame,
    ) -> dict[str, dict[str, Any]]:
        if isinstance(data, dict):
            normalized: dict[str, dict[str, Any]] = {}
            for name, yearly in data.items():
                if not isinstance(yearly, dict):
                    continue
                normalized[str(name).strip()] = {str(year).strip(): value for year, value in yearly.items()}
            return normalized

        if not isinstance(data, pd.DataFrame):
            raise ExcelWriterError(f"지원하지 않는 data 형식: {type(data)}")
        if data.empty:
            return {}

        name_col = None
        for candidate in ("표준계정", "계정명", "부문명"):
            if candidate in data.columns:
                name_col = candidate
                break
        if name_col is None:
            name_col = str(data.columns[0])

        year_cols = [
            str(col)
            for col in data.columns
            if str(col).isdigit() and len(str(col)) == 4
        ]
        if not year_cols:
            year_cols = [
                str(col)
                for col in data.columns
                if str(col) not in {name_col, "계정ID", "계정상세", "note_type", "note_title"}
            ]

        normalized: dict[str, dict[str, Any]] = {}
        for _, row in data.iterrows():
            name = str(row.get(name_col, "")).strip()
            if not name or name.lower() == "nan":
                continue
            normalized[name] = {year: row.get(year) for year in year_cols}
        return normalized

    def _read_id_row_map(self, ws: Worksheet) -> dict[str, int]:
        id_row_map: dict[str, int] = {}
        for row in ws.iter_rows(min_col=self.account_id_col, max_col=self.account_id_col):
            cell = row[0]
            if cell.value is None:
                continue
            value = str(cell.value).strip()
            if value.startswith(("ifrs", "dart_", "entity_")):
                id_row_map[value] = cell.row
        return id_row_map

    def _read_account_name_row_map(self, ws: Worksheet) -> dict[str, int]:
        row_map: dict[str, int] = {}
        for row in ws.iter_rows(min_col=self.account_name_col, max_col=self.account_name_col):
            cell = row[0]
            if cell.value is None:
                continue
            value = str(cell.value).strip()
            if not value:
                continue
            row_map[value] = cell.row
        return row_map

    def _read_year_col_map(self, ws: Worksheet) -> dict[str, int]:
        col_map: dict[str, int] = {}
        for cell in ws[self.year_header_row]:
            if cell.value is None:
                continue
            year = str(cell.value).strip()
            if year:
                col_map[year] = cell.column
        return col_map

    def _ensure_year_headers(self, ws: Worksheet, years: list[str]) -> dict[str, int]:
        col_map = self._read_year_col_map(ws)
        normalized_years = [str(y).strip() for y in years if str(y).strip()]
        if not normalized_years:
            return col_map

        next_col = max(col_map.values(), default=2) + 1
        for year in sorted(set(normalized_years)):
            if year in col_map:
                continue
            ws.cell(row=self.year_header_row, column=next_col, value=year)
            col_map[year] = next_col
            next_col += 1
        return col_map

    def _bootstrap_statement_rows(self, ws: Worksheet, df: pd.DataFrame) -> dict[str, int]:
        row_map: dict[str, int] = {}
        start_row = max(self.year_header_row + 2, 5)
        next_row = start_row

        for _, row in df.iterrows():
            account_id = str(row.get("계정ID", "")).strip()
            account_name = str(row.get("계정명", "")).strip()
            if not account_id or account_id in row_map:
                continue
            ws.cell(row=next_row, column=self.account_id_col, value=account_id)
            if account_name:
                ws.cell(row=next_row, column=self.account_name_col, value=account_name)
            row_map[account_id] = next_row
            next_row += 1
        return row_map

    def _bootstrap_named_rows(self, ws: Worksheet, names: list[str], start_row: int = 5) -> dict[str, int]:
        row_map: dict[str, int] = {}
        next_row = start_row
        for name in names:
            normalized = str(name).strip()
            if not normalized or normalized in row_map:
                continue
            ws.cell(row=next_row, column=self.account_name_col, value=normalized)
            row_map[normalized] = next_row
            next_row += 1
        return row_map

    def _get_sheet(self, sheet_name: str) -> Worksheet:
        if sheet_name not in self.wb.sheetnames:
            raise ExcelWriterError(f"시트를 찾을 수 없습니다: {sheet_name}")
        return self.wb[sheet_name]

    @staticmethod
    def _to_number_or_none(value: Any) -> float | int | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            if pd.isna(value):
                return None
            return value

        text = str(value).strip()
        if text in {"", "None", "nan", "NaN", "-"}:
            return None

        negative = text.startswith("(") and text.endswith(")")
        if negative:
            text = text[1:-1]

        cleaned = (
            text.replace(",", "")
            .replace(" ", "")
            .replace("\u00A0", "")
            .replace("원", "")
            .replace("백만원", "")
        )

        try:
            number = float(cleaned)
        except ValueError:
            return None
        if negative:
            number = -number
        return int(number) if number.is_integer() else number

    def save(self, output_path: str | None = None) -> Path:
        if output_path is None:
            target = self.template_path
        else:
            target = Path(output_path)
            target.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(target)
        return target
